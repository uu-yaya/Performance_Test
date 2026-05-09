import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_WORKBOOK = Path("D:/desktop_pet/桌宠类产品性能测试记录表.xlsx")
DEFAULT_RUNS_DIR = Path("D:/desktop_pet/perf_runs")


def to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip().replace("%", ""))
    except ValueError:
        return None


def percentile(values, p):
    clean = sorted(v for v in values if v is not None)
    if not clean:
        return None
    if len(clean) == 1:
        return clean[0]
    rank = (p / 100) * (len(clean) - 1)
    low = int(rank)
    high = min(low + 1, len(clean) - 1)
    weight = rank - low
    return clean[low] * (1 - weight) + clean[high] * weight


def avg(values):
    clean = [v for v in values if v is not None]
    return mean(clean) if clean else None


def max_or_none(values):
    clean = [v for v in values if v is not None]
    return max(clean) if clean else None


def round_or_blank(value, digits=3):
    if value is None:
        return ""
    return round(value, digits)


def read_csv(path):
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def find_summary_files(runs_dir):
    return sorted(Path(runs_dir).glob("**/summary.csv"))


def first_existing_key(row, candidates):
    lower_map = {k.lower(): k for k in row.keys()}
    for candidate in candidates:
        if candidate in row:
            return row[candidate]
        key = lower_map.get(candidate.lower())
        if key:
            return row[key]
    return None


def parse_presentmon_csv(path):
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    rows = read_csv(path)
    fps_values = []
    frame_ms = []
    gpu_busy = []
    for row in rows:
        fps = first_existing_key(row, ["FPS", "MsBetweenPresents", "msBetweenPresents"])
        if fps is not None and "Between" not in str(fps):
            fps_values.append(to_float(fps))
        ms = first_existing_key(row, ["MsBetweenPresents", "msBetweenPresents", "FrameTime", "Frame Time"])
        ms_value = to_float(ms)
        if ms_value:
            frame_ms.append(ms_value)
            if fps is not None and "Between" in str(fps):
                fps_values.append(1000 / ms_value)
        busy = first_existing_key(row, ["GPUBusy", "GPU Busy", "GPUBusy(ms)", "MsUntilDisplayed"])
        gpu_busy.append(to_float(busy))
    return {
        "fps_avg": avg(fps_values),
        "fps_1pct_low": percentile(fps_values, 1),
        "frametime_p95_ms": percentile(frame_ms, 95),
        "gpu_busy_avg": avg(gpu_busy),
        "presentmon_csv": str(path),
    }


def parse_hwinfo_csv(path):
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    rows = read_csv(path)
    power = []
    temp = []
    gpu_temp = []
    for row in rows:
        for key, value in row.items():
            key_l = key.lower()
            number = to_float(value)
            if number is None:
                continue
            if ("power" in key_l or "功耗" in key_l) and ("w" in key_l or "瓦" in key_l):
                power.append(number)
            if ("temperature" in key_l or "temp" in key_l or "温度" in key_l) and ("gpu" not in key_l):
                temp.append(number)
            if ("gpu" in key_l) and ("temperature" in key_l or "temp" in key_l or "温度" in key_l):
                gpu_temp.append(number)
    all_temp = gpu_temp or temp
    return {
        "power_avg_w": avg(power),
        "power_max_w": max_or_none(power),
        "temperature_max_c": max_or_none(all_temp),
        "hwinfo_csv": str(path),
    }


def ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row < 1:
            ws.append(headers)
    else:
        ws = wb.create_sheet(name)
        ws.append(headers)
    for idx, header in enumerate(headers, 1):
        ws.cell(1, idx).value = header
    return ws


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    border = Border(
        left=Side(style="thin", color="B8CDD1"),
        right=Side(style="thin", color="B8CDD1"),
        top=Side(style="thin", color="B8CDD1"),
        bottom=Side(style="thin", color="B8CDD1"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def ensure_instruction_sheet(wb):
    name = "自动采集说明"
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)

    rows = [
        ["用途", "说明"],
        ["一键采集", "使用 tools/collect_desktop_pet_perf.ps1 按产品、进程名、场景和时长自动采集 CPU、GPU、内存、显存、句柄、线程、IO、TCP 连接数。"],
        ["配置运行", "复制 tools/perf_collect_config.example.json，改产品名、进程名和场景后，用 tools/run_perf_from_config.ps1 批量执行。"],
        ["导入 Excel", "采集结束后运行 tools/import_perf_results.py，它会把 perf_runs 下的 summary.csv 和 raw_samples.csv 导入本工作簿。"],
        ["FPS/帧时间", "使用 PresentMon 或 NVIDIA FrameView 单独导出 CSV，再通过 import_perf_results.py 的 --presentmon-csv 参数导入。"],
        ["功耗/温度", "使用 HWiNFO 传感器日志导出 CSV，再通过 import_perf_results.py 的 --hwinfo-csv 参数导入。"],
        ["推荐命令", "powershell -ExecutionPolicy Bypass -File D:\\desktop_pet\\tools\\collect_desktop_pet_perf.ps1 -ProductId P001 -ProductName BongoCat -ProcessName BongoCat -Scenario 空闲常驻 -DurationSec 600"],
        ["导入命令", "C:\\Users\\zoeuuliu\\.cache\\codex-runtimes\\codex-primary-runtime\\dependencies\\python\\python.exe D:\\desktop_pet\\tools\\import_perf_results.py"],
        ["注意", "GPU/显存依赖 Windows GPU 性能计数器；FPS、1% Low、帧时间、功耗、温度需要外部工具日志，普通进程 API 无法稳定直接提供。"],
    ]
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110
    return ws


def append_unique(ws, key_col, rows):
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
      if row and row[key_col - 1]:
          existing.add(str(row[key_col - 1]))
    appended = 0
    for row in rows:
        key = str(row[key_col - 1])
        if key in existing:
            continue
        ws.append(row)
        existing.add(key)
        appended += 1
    return appended


def import_results(workbook_path, runs_dir, presentmon_csv=None, hwinfo_csv=None):
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path)
    ensure_instruction_sheet(wb)

    scenario_headers = [
        "记录ID", "产品ID", "产品名称", "场景", "轮次", "开始时间", "结束时间", "时长分钟",
        "CPU平均%", "CPU P95%", "CPU最大%", "GPU平均%", "GPU P95%", "GPU最大%",
        "内存平均MB", "内存最大MB", "显存平均MB", "显存最大MB", "句柄增长",
        "线程增长", "FPS平均", "1% Low", "帧时间P95ms", "功耗W", "最高温度C",
        "体验评分1-5", "风险等级", "现象/备注"
    ]
    samples_headers = [
        "样本ID", "记录ID", "时间戳", "进程名", "CPU%", "GPU%", "内存MB", "显存MB",
        "句柄数", "线程数", "FPS", "帧时间ms", "GPU Busy%", "功耗W", "温度C", "备注"
    ]
    auto_headers = [
        "导入ID", "产品ID", "产品名称", "场景", "开始时间", "采样数", "原始CSV",
        "PresentMon/FrameView CSV", "HWiNFO CSV", "导入时间"
    ]

    scenario_ws = ensure_sheet(wb, "场景记录", scenario_headers)
    samples_ws = ensure_sheet(wb, "采样数据", samples_headers)
    auto_ws = ensure_sheet(wb, "自动采集导入", auto_headers)

    present = parse_presentmon_csv(presentmon_csv)
    hwinfo = parse_hwinfo_csv(hwinfo_csv)

    scenario_rows = []
    sample_rows = []
    auto_rows = []

    for summary_file in find_summary_files(runs_dir):
        summary_rows = read_csv(summary_file)
        if not summary_rows:
            continue
        summary = summary_rows[0]
        run_dir = summary_file.parent
        raw_csv = Path(summary.get("raw_csv") or run_dir / "raw_samples.csv")
        raw_rows = read_csv(raw_csv) if raw_csv.exists() else []
        product_id = summary.get("product_id", "")
        product_name = summary.get("product_name", "")
        scenario = summary.get("scenario", "")
        started_at = summary.get("started_at", "")
        record_id = f"AUTO-{run_dir.name}"

        scenario_rows.append([
            record_id,
            product_id,
            product_name,
            scenario,
            1,
            started_at,
            "",
            round_or_blank(to_float(summary.get("duration_sec")) / 60 if summary.get("duration_sec") else None, 2),
            round_or_blank(to_float(summary.get("cpu_avg_percent"))),
            round_or_blank(to_float(summary.get("cpu_p95_percent"))),
            round_or_blank(to_float(summary.get("cpu_max_percent"))),
            round_or_blank(to_float(summary.get("gpu_avg_percent"))),
            round_or_blank(to_float(summary.get("gpu_p95_percent"))),
            round_or_blank(to_float(summary.get("gpu_max_percent"))),
            round_or_blank(to_float(summary.get("memory_avg_mb"))),
            round_or_blank(to_float(summary.get("memory_max_mb"))),
            round_or_blank(to_float(summary.get("dedicated_vram_avg_mb"))),
            round_or_blank(to_float(summary.get("dedicated_vram_max_mb"))),
            round_or_blank(to_float(summary.get("handles_delta"))),
            round_or_blank(to_float(summary.get("threads_delta"))),
            round_or_blank(present.get("fps_avg")),
            round_or_blank(present.get("fps_1pct_low")),
            round_or_blank(present.get("frametime_p95_ms")),
            round_or_blank(hwinfo.get("power_avg_w")),
            round_or_blank(hwinfo.get("temperature_max_c")),
            "",
            "",
            summary.get("notes", ""),
        ])

        for idx, row in enumerate(raw_rows, 1):
            sample_rows.append([
                f"{record_id}-S{idx:05d}",
                record_id,
                row.get("timestamp", ""),
                row.get("process_details", ""),
                round_or_blank(to_float(row.get("cpu_percent"))),
                round_or_blank(to_float(row.get("gpu_percent"))),
                round_or_blank(to_float(row.get("memory_working_set_mb"))),
                round_or_blank(to_float(row.get("dedicated_vram_mb"))),
                round_or_blank(to_float(row.get("handles"))),
                round_or_blank(to_float(row.get("threads"))),
                "",
                "",
                "",
                "",
                "",
                row.get("gpu_engine_types", ""),
            ])

        auto_rows.append([
            record_id,
            product_id,
            product_name,
            scenario,
            started_at,
            summary.get("sample_count", ""),
            str(raw_csv),
            present.get("presentmon_csv", ""),
            hwinfo.get("hwinfo_csv", ""),
            datetime.now().isoformat(timespec="seconds"),
        ])

    appended_scenarios = append_unique(scenario_ws, 1, scenario_rows)
    appended_samples = append_unique(samples_ws, 1, sample_rows)
    appended_auto = append_unique(auto_ws, 1, auto_rows)

    for ws in (scenario_ws, samples_ws, auto_ws):
        style_sheet(ws)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max(len(str(ws.cell(1, col).value)) + 4, 12), 36)

    wb.save(workbook_path)
    return appended_scenarios, appended_samples, appended_auto


def main():
    parser = argparse.ArgumentParser(description="Import desktop pet performance run CSVs into the Excel tracker.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--runs-dir", default=str(DEFAULT_RUNS_DIR))
    parser.add_argument("--presentmon-csv", default="")
    parser.add_argument("--hwinfo-csv", default="")
    args = parser.parse_args()

    result = import_results(args.workbook, args.runs_dir, args.presentmon_csv, args.hwinfo_csv)
    print(f"Imported scenario rows: {result[0]}")
    print(f"Imported sample rows: {result[1]}")
    print(f"Imported import-log rows: {result[2]}")


if __name__ == "__main__":
    main()
